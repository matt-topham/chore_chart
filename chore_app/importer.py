from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .db import connect_db
from .scheduler import initial_due_date, normalize_frequency

TASK_CORRECTIONS = {
    "Tidy Bathrrom": "Tidy Bathroom",
}

AREA_CORRECTIONS = {
    "Vacuum Mattress": "Bedroom",
}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def import_workbook(
    path: Path,
    database_path: str | Path,
    replace: bool = False,
    imported_on: date | None = None,
) -> dict[str, int]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_clean_text(value) for value in next(rows)]
    header_map = {name.lower(): index for index, name in enumerate(headers)}

    required = {"task", "area", "frequency", "preferred day"}
    missing = sorted(required - set(header_map))
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}")

    imported_on = imported_on or date.today()
    frequency_counts: dict[str, int] = defaultdict(int)
    inserted = 0
    updated = 0
    skipped = 0

    with connect_db(database_path) as db:
        if replace:
            db.execute("DELETE FROM completions")
            db.execute("DELETE FROM chores")

        for source_row, values in enumerate(rows, start=2):
            task = _clean_text(values[header_map["task"]])
            if not task:
                skipped += 1
                continue

            task = TASK_CORRECTIONS.get(task, task)
            area = _clean_text(values[header_map["area"]]) or "General"
            area = AREA_CORRECTIONS.get(task, area)
            frequency = normalize_frequency(_clean_text(values[header_map["frequency"]]))
            preferred_day = _clean_text(values[header_map["preferred day"]])
            notes_index = header_map.get("notes/supplies")
            notes = _clean_text(values[notes_index]) if notes_index is not None else ""

            group_index = frequency_counts[frequency]
            frequency_counts[frequency] += 1
            first_due = initial_due_date(
                frequency,
                preferred_day,
                imported_on,
                group_index,
            ).isoformat()

            existing = db.execute(
                "SELECT id FROM chores WHERE task = ? AND area = ?",
                (task, area),
            ).fetchone()

            if existing:
                db.execute(
                    """
                    UPDATE chores
                    SET frequency = ?, preferred_day = ?, notes = ?, source_row = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (frequency, preferred_day, notes, source_row, existing["id"]),
                )
                updated += 1
            else:
                db.execute(
                    """
                    INSERT INTO chores
                        (task, area, frequency, preferred_day, notes, first_due, source_row)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (task, area, frequency, preferred_day, notes, first_due, source_row),
                )
                inserted += 1

        db.commit()

    return {"inserted": inserted, "updated": updated, "skipped": skipped}
